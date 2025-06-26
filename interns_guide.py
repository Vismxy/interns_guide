from flask import Flask, render_template_string, request, redirect, url_for
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# In-memory intern data
interns = [
    {"name": "Rahul", "dept": "Python", "status": "Completed"},
    {"name": "Sneha", "dept": "A360", "status": "In Progress"},
    {"name": "Vikram", "dept": "Python", "status": "Completed"},
]

EXCEL_FILE = "interns_data.xlsx"

def save_to_excel(name, email, dept, status="In Progress"):
    from pathlib import Path
    filepath = Path(__file__).parent / EXCEL_FILE
    print("Saving Excel to:", filepath)  # <-- This shows the actual location

    if not filepath.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Interns"
        ws.append(["Name", "Email", "Department", "Status"])
    else:
        wb = load_workbook(filepath)
        ws = wb["Interns"]

    ws.append([name, email, dept, status])
    print("Intern details saved to:", filepath)

    wb.save(filepath)


base_template = """
<!DOCTYPE html>
<html lang='en'>
<head>
    <meta charset='UTF-8'>
    <title>Interns Guide - {{ title }}</title>
    <style>
        body {
            margin: 0;
            font-family: 'Segoe UI', sans-serif;
            background: linear-gradient(to right, #e0f7fa, #fce4ec);
        }
        .container {
            display: flex;
        }
        .sidebar {
            width: 220px;
            background-color: #3f51b5;
            height: 100vh;
            padding-top: 20px;
            position: fixed;
        }
        .sidebar h1 {
            color: #ffffff;
            text-align: center;
            font-size: 22px;
            margin-bottom: 30px;
        }
        .sidebar a {
            display: block;
            color: white;
            padding: 12px 20px;
            text-decoration: none;
            font-weight: bold;
        }
        .sidebar a:hover {
            background-color: #5c6bc0;
        }
        .main-content {
            margin-left: 220px;
            padding: 30px;
            width: calc(100% - 220px);
        }
        .card {
            background: white;
            padding: 25px;
            border-radius: 12px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
            max-width: 900px;
            margin: auto;
        }
        h2 {
            color: #003366;
            border-bottom: 2px solid #003366;
            padding-bottom: 8px;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
        }
        th, td {
            padding: 10px;
            border: 1px solid #ccc;
            text-align: center;
        }
        input[type='text'], input[type='email'] {
            width: 95%;
            padding: 10px;
            margin: 6px 0;
            border-radius: 5px;
            border: 1px solid #ccc;
        }
        input[type='submit'] {
            padding: 10px 20px;
            background-color: #3f51b5;
            color: white;
            border: none;
            border-radius: 5px;
            cursor: pointer;
        }
        input[type='submit']:hover {
            background-color: #5c6bc0;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="sidebar">
            <h1>Interns Guide</h1>
            <a href='/'>Home</a>
            <a href='/registration'>Registration</a>
            <a href='/onboarding'>Onboarding</a>
            <a href='/modules'>Modules</a>
            <a href='/schedule'>Schedule</a>
            <a href='/resources'>Resources</a>
            <a href='/assessments'>Assessments</a>
            <a href='/contact'>Contact</a>
        </div>

        <div class="main-content">
            <div class="card">
                <h2>{{ title }}</h2>
                {{ content | safe }}
            </div>
        </div>
    </div>
</body>
</html>
"""

@app.route('/')
def home():
    total = len(interns)
    completed = sum(1 for i in interns if i['status'].lower() == "completed")
    in_progress = total - completed
    content = f"""
        <p><strong>Total Trained Interns:</strong> {total}</p>
        <p><strong>In Progress:</strong> {in_progress}</p>
        <p><strong>Completed:</strong> {completed}</p>
    """
    return render_template_string(base_template, title="Dashboard", content=content)

@app.route('/registration', methods=['GET', 'POST'])
def registration():
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        dept = request.form.get('dept')
        interns.append({"name": name, "dept": dept, "status": "In Progress"})
        save_to_excel(name, email, dept)
        return redirect(url_for('home'))

    content = """
        <form method='post'>
            Name:<br><input type='text' name='name' required><br><br>
            Email:<br><input type='email' name='email' required><br><br>
            Department:<br><input type='text' name='dept' required><br><br>
            <input type='submit' value='Register'>
        </form>
    """
    return render_template_string(base_template, title="Registration Form", content=content)

@app.route('/onboarding')
def onboarding():
    rows = "".join(f"<tr><td>{i['name']}</td><td>{i['dept']}</td><td>{i['status']}</td></tr>" for i in interns)
    content = f"""
        <table>
            <tr><th>Name</th><th>Department</th><th>Status</th></tr>
            {rows}
        </table>
    """
    return render_template_string(base_template, title="Onboarding - Intern List", content=content)

@app.route('/modules')
def modules():
    content = """
        <ul>
            <li><a href='https://www.automationanywhere.com/products/robotic-process-automation' target='_blank'>A360 Learning Link</a></li>
            <li><a href='https://www.learnpython.org/' target='_blank'>Python Free Learning</a></li>
        </ul>
    """
    return render_template_string(base_template, title="Modules", content=content)

@app.route('/schedule')
def schedule():
    topics = [
        "Learning Basics of A360",
        "Getting Familiar with A360 Interface",
        "Creating Your First Bot in A360",
        "Understanding Recorder & Packages",
        "Using Variables and Logic in A360",
        "Deploying and Scheduling Bots",
        "Python Basics: Syntax and Variables",
        "Control Structures in Python",
        "Functions and Modules in Python",
        "Python Automation with A360 Integration"
    ]
    content = "<table><tr><th>Day</th><th>Topic</th></tr>"
    for i in range(1, 11):
        content += f"<tr><td>Day {i}</td><td>{topics[i-1]}</td></tr>"
    content += "</table>"
    return render_template_string(base_template, title="10-Day Training Schedule", content=content)

@app.route('/resources')
def resources():
    content = "<p>Resources will be added here based on module requirements.</p>"
    return render_template_string(base_template, title="Resources", content=content)

@app.route('/assessments')
def assessments():
    content = """
    <p>
    1. Shopping list project (added all the items required in a cart)<br>
    2. Customer Onboarding Challenge<br>
    3. Conditional Salary Increase Challenge<br>
    4. Employee details extracting<br>
    5. Information Extraction (From Bonafide)<br>
    6. Information Extraction (From a folder containing more than 2 files)
    </p>
    """
    return render_template_string(base_template, title="Assessments", content=content)

@app.route('/contact')
def contact():
    content = "<p>Priyadharsini - 90786XXXXX</p><p>Yuvaraj - 89709XXXXX</p><p>Rajesh - 94563XXXXX</p>"
    return render_template_string(base_template, title="Contact", content=content)

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 10000))
    app.run(debug=False, host='0.0.0.0', port=port)
